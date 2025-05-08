import os
import re
import json
import pdfplumber
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook, load_workbook

class FacturadorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Procesador de Facturas HAASoluciones 2025")
        self.root.geometry("500x300")
        
        # Configuración
        self.config_file = "config.json"
        self.campos_requeridos = ["Fecha", "Proveedor", "CUIT", "N° Factura", "Importe"]
        self.cargar_config()
        
        # Patrones optimizados
        self.patrones = {
            'proveedor': r"Razón Social:\s*([^\n]+)",
            'cuit': r"CUIT:\s*(\d{11})",
            'fecha': r"Fecha de Emisión:\s*(\d{2}/\d{2}/\d{4})",
            'nro_factura': r"Comp\. Nro:\s*(\d+)",
            'importe': r"Importe Total:\s*\$?\s*([\d.,]+)"
        }
        
        self.setup_ui()
    
    def cargar_config(self):
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r') as f:
                    config = json.load(f)
                    self.output_dir = config.get('output_dir', os.path.expanduser("~"))
                    self.last_processed = config.get('last_processed', 0)
            else:
                self.output_dir = os.path.expanduser("~")
                self.last_processed = 0
        except:
            self.output_dir = os.path.expanduser("~")
            self.last_processed = 0
    
    def guardar_config(self):
        config = {
            'output_dir': self.output_dir,
            'last_processed': self.last_processed
        }
        with open(self.config_file, 'w') as f:
            json.dump(config, f)
    
    def setup_ui(self):
        main_frame = tk.Frame(self.root, padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Botón para procesar carpeta
        tk.Button(
            main_frame, 
            text="PROCESAR CARPETA DE FACTURAS", 
            command=self.iniciar_proceso,
            bg="#4CAF50", 
            fg="white",
            height=2,
            font=('Arial', 10, 'bold')
        ).pack(fill=tk.X, pady=10)
        
        # Barra de progreso
        self.progress = ttk.Progressbar(main_frame, orient="horizontal", mode="determinate")
        self.progress.pack(fill=tk.X, pady=5)
        
        # Etiqueta de estado
        self.status_label = tk.Label(main_frame, text="Listo para procesar", fg="gray")
        self.status_label.pack()
        
        # Footer
        tk.Label(
            main_frame, 
            text="HAASoluciones 2025 - Analista de Sistemas", 
            fg="gray"
        ).pack(side=tk.BOTTOM)
    
    def iniciar_proceso(self):
        # Seleccionar carpeta de salida
        nueva_ruta = filedialog.askdirectory(
            title="Seleccionar carpeta para guardar el archivo Excel",
            initialdir=self.output_dir
        )
        
        if nueva_ruta:
            self.output_dir = nueva_ruta
            self.guardar_config()
            self.procesar_carpeta()
    
    def procesar_carpeta(self):
        carpeta = filedialog.askdirectory(
            title="Seleccionar carpeta con facturas PDF",
            initialdir=self.output_dir
        )
        
        if not carpeta:
            return
        
        archivos = [f for f in os.listdir(carpeta) if f.lower().endswith('.pdf')]
        if not archivos:
            messagebox.showwarning("Advertencia", "No se encontraron archivos PDF en la carpeta seleccionada.")
            return
        
        self.status_label.config(text=f"Procesando {len(archivos)} archivos...", fg="blue")
        self.progress['maximum'] = len(archivos)
        self.progress['value'] = 0
        self.root.update()
        
        datos = []
        
        for i, archivo in enumerate(archivos, 1):
            try:
                resultado = self.procesar_archivo_pdf(os.path.join(carpeta, archivo))
                if resultado:
                    datos.append(resultado)
            except Exception:
                continue
            
            self.progress['value'] = i
            self.status_label.config(text=f"Procesando {i} de {len(archivos)} archivos...")
            self.root.update()
        
        if datos:
            archivo_excel = os.path.join(self.output_dir, "facturas.xlsx")
            self.guardar_excel(datos, archivo_excel)
            self.mostrar_resultado(len(datos), archivo_excel)
        else:
            messagebox.showwarning(
                "Advertencia", 
                "No se encontraron facturas válidas en los archivos procesados."
            )
        
        self.status_label.config(text="Proceso finalizado", fg="green")
        self.progress['value'] = 0
    
    def procesar_archivo_pdf(self, ruta_pdf):
        with pdfplumber.open(ruta_pdf) as pdf:
            texto = "\n".join(page.extract_text() for page in pdf.pages if page.extract_text())
        
        # Extracción con limpieza
        proveedor = self.extraer_dato(texto, 'proveedor')
        cuit = self.extraer_dato(texto, 'cuit')
        fecha = self.extraer_dato(texto, 'fecha')
        nro_factura = self.extraer_dato(texto, 'nro_factura')
        
        # Manejo especial del importe
        importe_str = self.extraer_dato(texto, 'importe')
        importe = self.formatear_importe(importe_str) if importe_str else "$0.00"
        
        # Validación
        if not all([proveedor, cuit, fecha, nro_factura, importe != "$0.00"]):
            raise ValueError("Datos incompletos")
        
        return [fecha, proveedor, cuit, nro_factura, importe]
    
    def extraer_dato(self, texto, clave):
        try:
            match = re.search(self.patrones[clave], texto, re.IGNORECASE)
            if not match:
                return ""
            
            dato = match.group(1).strip()
            
            # Limpieza específica para proveedor
            if clave == 'proveedor':
                dato = re.sub(r'\s*Fecha de Emisión.*$', '', dato).strip()
            
            return dato
        except Exception:
            return ""
    
    def formatear_importe(self, importe_str):
        try:
            valor = float(importe_str.replace(".", "").replace(",", "."))
            return f"${valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except:
            return "$0.00"
    
    def guardar_excel(self, datos, ruta_archivo):
        if os.path.exists(ruta_archivo):
            wb = load_workbook(ruta_archivo)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(self.campos_requeridos)
        
        for fila in datos:
            ws.append(fila)
        
        # Ajustar columnas
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)
        
        wb.save(ruta_archivo)
    
    def mostrar_resultado(self, cantidad, ruta_archivo):
        respuesta = messagebox.askyesno(
            "Proceso completado",
            f"Se procesaron {cantidad} facturas correctamente.\n\n"
            f"Archivo guardado en:\n{ruta_archivo}\n\n"
            "¿Desea realizar otra operación?",
            icon='info'
        )
        
        if not respuesta:
            messagebox.showinfo(
                "Gracias",
                "Gracias Walter por no confiar en la tecnología\n\n¡Saludos!",
                icon='info'
            )
            self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = FacturadorApp(root)
    root.mainloop()