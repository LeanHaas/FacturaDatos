import os
import re
import pdfplumber
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook

# Ruta del archivo Excel
archivo_excel = "facturas.xlsx"

# Función para extraer datos desde un PDF
def extraer_datos(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        texto = ""
        for pagina in pdf.pages:
            contenido = pagina.extract_text()
            if contenido:
                texto += contenido + "\n"

    proveedor = ""
    cuit = ""
    fecha = ""
    nro_factura = ""
    monto = 0.0

    # Buscar CUIT (toma el primero que aparece)
    cuit_match = re.search(r'\bCUIT[:\s]*([203347]{2}\d{8})\b', texto)
    if cuit_match:
        cuit = cuit_match.group(1)

    # Buscar fecha de emisión
    fecha_match = re.search(r'Fecha de Emisión[:\s]*([0-3]?\d/[01]?\d/\d{4})', texto)
    if fecha_match:
        fecha = fecha_match.group(1)
    else:
        # fallback: buscar cualquier fecha como dd/mm/aaaa
        fecha_gen = re.search(r'\b([0-3]?\d/[01]?\d/\d{4})\b', texto)
        if fecha_gen:
            fecha = fecha_gen.group(1)

    # Buscar número de factura
    nro_factura_match = re.search(r'Comp\.? Nro[:\s]*(\d{4}\s*\d{8})', texto)
    if nro_factura_match:
        nro_factura = nro_factura_match.group(1).replace(" ", "-")

    # Buscar proveedor
    proveedor_match = re.search(r'Apellido y Nombre / Razón Social[:\s]*(.+)', texto)
    if proveedor_match:
        proveedor = proveedor_match.group(1).strip()
    else:
        # Intentar capturar nombre del emisor (como en tu modelo: HAAS LEANDRO DAMIAN)
        proveedor_candidatos = re.findall(r'(?i)(?:^|\n)([A-ZÁÉÍÓÚÑ ]{3,})\n', texto)
        if proveedor_candidatos:
            proveedor = proveedor_candidatos[0].strip().title()

    # Buscar monto total
    monto_match = re.search(r'Importe Total[:\s]*\$?\s*([\d\.,]+)', texto)
    if monto_match:
        monto_str = monto_match.group(1).replace(".", "").replace(",", ".")
        try:
            monto = float(monto_str)
        except:
            monto = 0.0

    # Validar datos mínimos
    if not any([proveedor, cuit, fecha, nro_factura, monto]):
        raise ValueError("No se pudieron extraer datos relevantes del archivo.")

    # Formatear monto
    monto_str = f"${monto:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    return [fecha, monto_str, cuit, proveedor, nro_factura]

# Función para guardar en Excel
def guardar_en_excel(datos_lista):
    if os.path.exists(archivo_excel):
        wb = load_workbook(archivo_excel)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Fecha", "Monto", "CUIT", "Proveedor", "Número de Factura"])

    for datos in datos_lista:
        ws.append(datos)
    wb.save(archivo_excel)

# Procesamiento masivo
def seleccionar_carpeta():
    carpeta = filedialog.askdirectory()
    if carpeta:
        datos_completos = []
        errores = []
        for archivo in os.listdir(carpeta):
            if archivo.lower().endswith(".pdf"):
                ruta_pdf = os.path.join(carpeta, archivo)
                try:
                    datos = extraer_datos(ruta_pdf)
                    datos_completos.append(datos)
                except Exception as e:
                    errores.append(f"{archivo}: {e}")
        if datos_completos:
            guardar_en_excel(datos_completos)
            messagebox.showinfo("Éxito", f"{len(datos_completos)} archivos procesados correctamente.")
        if errores:
            errores_txt = "\n".join(errores)
            with open("errores_factura.txt", "w", encoding="utf-8") as f:
                f.write(errores_txt)
            messagebox.showwarning("Errores detectados", f"{len(errores)} archivos con problemas. Ver 'errores_factura.txt'.")

# Interfaz gráfica
ventana = tk.Tk()
ventana.title("Procesador Masivo de Facturas PDF")
ventana.geometry("400x150")

btn_masivo = tk.Button(ventana, text="Seleccionar carpeta con PDFs", command=seleccionar_carpeta, height=3, width=35)
btn_masivo.pack(pady=40)

ventana.mainloop()
